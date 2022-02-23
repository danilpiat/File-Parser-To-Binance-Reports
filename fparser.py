import pandas as pd
from openpyxl import load_workbook
from file_read_backwards import FileReadBackwards
import os

def input_report(filename):
    # file = open(r''+filename+ '.txt', 'r' )
    # tempmas = []
    # delmas = []
    # count = 0
    # for num, line in enumerate(file, 0):
    #     if str('Coin') in line:
    #         count += 1
    #         if count==2:
    #             prevline=line
    #             tempmas.append(prevline)
    #             for line in file:
    #                 tempmas.append(line)
    #                 delmas.append(num)
    #                 num+=1
    #     elif str('Binance') in line:
    #         delmas.append(num)
    # file.close()
    # file = open(r'' + filename + '.txt', 'r')
    # lines = file.readlines()
    # for i in reversed(delmas):
    #     del lines[i]
    # file.close()
    # file = open(r''+filename+ '.txt', 'w' )
    # file.writelines(lines)
    # file.close()
    # outf = open(filename+'2'+'.txt', 'w+')
    # outf.writelines(tempmas)
    # outf.close()


    df_1 = pd.read_csv(filename+ '.txt', sep='\t')
    df_1.to_excel('Отчеты_part1.xlsx', 'Отчеты_part1', index=False)
    # df_2 = pd.read_csv(filename+'2'+'.txt', sep='\t')
    # df_2.to_excel('Отчеты_part2.xlsx', 'Отчеты_part2', index=False)

def extract_columns():
    cols_1 = ['Coin ', 'BuyDate ', 'CloseDate ', 'BuyPrice ', 'SellPrice ', 'Spent USDT ', 'Profit ',
              'ChannelName ', 'SellReason ', 'dBTC ', 'd24BTC ', 'dMarket ', 'dM24 ',
              'dBTC5m ', 'd24h ', 'd3h ', 'd15m ', 'd1m ', 'dBTC1m']
    # cols_2 = ['Coin ', 'BuyDate ', 'CloseDate ', 'BuyPrice ', 'SellPrice ', 'Spent USDT ', 'Profit ',
    #           'ChannelName ', 'SellReason ', 'dBTC ', 'd24BTC ', 'dMarket ', 'dM24 ',
    #           'dBTC5m ', 'd24h ', 'd3h ', 'd15m ', 'd1m ', 'dBTC1m ']
    excelfile_1 = pd.read_excel('Отчеты_part1.xlsx', sheet_name='Отчеты_part1', usecols=cols_1)
    excelfile_1.to_excel('Отчеты_part1_selection.xlsx', 'Отчеты_part1', index=False)
    # excelfile_2 = pd.read_excel('Отчеты_part2.xlsx', sheet_name='Отчеты_part2', usecols=cols_2)
    # excelfile_2.to_excel('Отчеты_part2_selection.xlsx', 'Отчеты_part2', index=False)

def copy_xls():
    wb1 = load_workbook("Отчеты_part1_selection.xlsx")
    ws1 = wb1.active
    # wb2 = load_workbook("Отчеты_part2_selection.xlsx")
    # ws2 = wb2.active
    mr = ws1.max_row
    mc = ws1.max_column
    # mr_2=ws2.max_row
    # mc_2=ws2.max_column
    #
    #
    # for i in range(1, 20):
    #     a=464
    #
    #     for j in range(2, 464):
    #         c = ws2.cell(row=j, column=i)
    #         for k in range(a, 926):
    #             ws1.cell(row=k, column=i).value = c.value
    #             a += 1
    #             break
    # wb1.save("Отчеты_part1_selection.xlsx")
    f = load_workbook("Отчеты_part1_selection.xlsx")
    df = f.active
    df.cell(row=1, column = 20).value = str("index")
    for i in range(0,mr-1):
        df.cell(row=i+2, column=20).value = i
    f.save("Отчеты_part1_selection.xlsx")

date_mas = []
time_mas = []
def exctract_date():
    selection_excel_buy_date = pd.ExcelFile('Отчеты_part1_selection.xlsx')
    df = selection_excel_buy_date.parse("Отчеты_part1")
    buydate_list=df['BuyDate '].tolist()
    date_list = []
    date_mas_same = []
    for i in range(len(buydate_list)):
        date_mas.append(buydate_list[i].split())
    for j in range(len(date_mas)):
        temp_mas = date_mas[j]
        time_mas.append(temp_mas[1][0:5])
    # unique_date(date_mas_same)
    selection_excel_buy_date.close()

coin_list = []
def exctract_coin():
    selection_excel_coin = pd.ExcelFile('Отчеты_part1_selection.xlsx')
    df = selection_excel_coin.parse("Отчеты_part1")
    coin_temp_list = df['Coin '].tolist()
    coin_temp_list = [x.strip(' ') for x in coin_temp_list]
    for i in range(len(coin_temp_list)):
        coin_list.append(coin_temp_list[i])
    selection_excel_coin.close()
    # counter = {}
    # for elem in coin_list:
    #     counter[elem] = counter.get(elem, 0) + 1
    # doubles = {element: count for element, count in counter.items() if count > 1}
    # print(doubles)

def search_in_log(logs_directory):
    exctract_date()
    exctract_coin()
    filename_mas = []
    lines_mas = []
    for element in os.scandir(logs_directory):
        filename_mas.append(element.name)

    for i in range(len(date_mas)):
        for j in range(len(filename_mas)):
            if filename_mas[j].find(date_mas[i][0]) == 4:
                f = FileReadBackwards(logs_directory + '' + filename_mas[j], encoding="utf-8")
                current_line=[]
                for line in f:
                    previous_line = current_line
                    current_line = line
                    if date_mas[i][1][0:7] in line:
                        if str('Signal USDT-'+ coin_list[i]) in line:
                            lines_mas.append(line)
                            lines_mas.append(previous_line)
                            lines_mas.append(i)
                        elif str('Starting new MoonShot market: USDT-'+ coin_list[i]) in line:
                            lines_mas.append(line)
                            lines_mas.append(previous_line)
                            lines_mas.append(i)

    with open('out.txt', 'w') as f:
        f.writelines(f'{row}\n' for row in lines_mas)
    f.close()

hvol = []
h3vol = []
BTC72h = []
PumpQ = []
sellX2 = []
PumpD = []
PumpsCount = []
SellProbe = []
indexline = []
def formatting():
    firstline = []
    secondline = []
    with open('out.txt', 'r') as f:
        for line in f:
            if str("Signal") in line:
                firstline.append(line)
            elif str("PumpQ") in line:
                secondline.append(line)
            elif str('Starting') in line:
                firstline.append(line)
            else:
                indexline.append(int(line))
    # BTC72h.append(str('72hBTC'))
    # counter = {}
    # for elem in indexline:
    #     counter[elem] = counter.get(elem, 0) + 1
    # doubles = {element: count for element, count in counter.items() if count > 1}
    # print(doubles)
    for i in range(len(firstline)):
        if firstline[i].find('72hBTC')!=-1:
            indexes = firstline[i].find('72hBTC')
            indexesend = firstline[i].find('dMarkets: ')
            BTC72h.append(firstline[i][indexes+7:indexesend])
        elif firstline[i].find('72hBTC')==-1:
            BTC72h.append('None')

    # hvol.append(str('VH1'))
    # h3vol.append(str('VH3'))
    # PumpQ.append(str('PumpQ'))
    # sellX2.append(str('sellX2'))
    # PumpD.append(str('PumpD'))
    # PumpsCount.append(str('PumpsCount'))
    # SellProbe.append(str('SellProbe'))
    for i in range(len(secondline)):
        if secondline[i].find('hvol')!=-1:
            indexes = secondline[i].find('hvol')
            indexesend = secondline[i].find('h3vol')
            hvol.append(secondline[i][indexes+5:indexesend])

            secondline[i].find('h3vol')
            indexes = secondline[i].find('h3vol')
            indexesend = secondline[i].find('sellX2')
            h3vol.append(secondline[i][indexes+6:indexesend])

            secondline[i].find('PumpQ')
            indexes = secondline[i].find('PumpQ')
            indexesend = secondline[i].find('24vol')
            PumpQ.append(secondline[i][indexes+6:indexesend])

            secondline[i].find('PumpD')
            indexes = secondline[i].find('PumpD')
            indexesend = secondline[i].find('PumpHDelta')
            PumpD.append(secondline[i][indexes+6:indexesend])

            secondline[i].find('sellX2')
            indexes = secondline[i].find('sellX2')
            indexesend = secondline[i].find('PumpsCount')
            sellX2.append(secondline[i][indexes+7:indexesend])

            secondline[i].find('PumpsCount')
            indexes = secondline[i].find('PumpsCount')
            indexesend = secondline[i].find('SellProb')
            PumpsCount.append(secondline[i][indexes+11:indexesend])

            secondline[i].find('SellProb')
            indexes = secondline[i].find('SellProb')
            indexesend = secondline[i].find('Delta24h')
            SellProbe.append(secondline[i][indexes+9:indexesend])
    f.close()

def insertintable():
    wb = load_workbook("Отчеты_part1_selection.xlsx")

    ws = wb.active
    indexcol = ws['T']
    hvolcol = ws['U']
    h3volcol = ws['V']
    BTC72hcol = ws['W']
    PumpQcol = ws['X']
    sellX2col = ws['Y']
    PumpDcol = ws['Z']
    PumpsCountcol = ws['AA']
    SellProbecol = ws['AB']

    # Print the contents
    count = 0
    for y in indexline:
        for x in range(len(indexcol)):
            if x=="indexes":
                x-=1
            else:
                if y==indexcol[x].value:
                    hvolcol[y].value = hvol[count]
                    h3volcol[y].value = h3vol[count]
                    BTC72hcol[y].value = BTC72h[count]
                    PumpQcol[y].value = PumpQ[count]
                    sellX2col[y].value = sellX2[count]
                    PumpDcol[y].value = PumpD[count]
                    PumpsCountcol[y].value = PumpsCount[count]
                    SellProbecol[y].value = SellProbe[count]
                    count+=1

    wb.save('newfile.xlsx')

def main():
     try:
         f = input("Введите путь до файла и имя файла в формате 'Диск:\Папка\Папка\Название_файла': \n")
         input_report(f)
     except OSError as e:
        print('Ошибка при открытии файла: ', str(e))
        return
     extract_columns()
     copy_xls()
     try:
         l = input("Введите путь до папки с логами в формате 'Диск:\Папка\Папка\\': \n")
         search_in_log(l)
     except OSError as e:
        print('Ошибка при открытии файла: ', str(e))
        return
     formatting()
     insertintable()

# VH1	VH3	72hBTC	PumpQ	sellX2	PumpD	PumpsCount	SellProbe


if __name__ == "__main__":
    main()